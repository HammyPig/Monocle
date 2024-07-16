import csv
from datetime import datetime
from openpyxl import load_workbook

class Portfolio:
    class Transaction:
        def __init__(self, date, action, stock, units, amount):
            self.date = date
            self.action = action
            self.stock = stock
            self.units = units
            self.amount = amount
        
        def __repr__(self):
            return f"{self.date}: {self.action} {self.units}x{self.stock} for {self.amount}"
        
    class Dividend:
        def __init__(self, date, amount):
            self.date = date
            self.amount = amount

    def __init__(self):
        self.transactions = []
        self.dividends = []

    def print_transactions(self):
        for t in self.transactions:
            print(t.date, t.action, t.stock, t.units, t.amount)

    def validate_transactions(self):
        shares = {}
        invalid_shares = {}

        for t in self.transactions:
            if t.stock not in shares: shares[t.stock] = 0

            if t.action == "buy":
                shares[t.stock] += t.units
            elif t.action == "sell":
                shares[t.stock] -= t.units
                if shares[t.stock] < 0:
                    if t.stock not in invalid_shares: invalid_shares[t.stock] = 0
                    invalid_shares[t.stock] += -shares[t.stock]
                    shares[t.stock] = 0
                    print(t)

        print(invalid_shares)

    def sort_transactions_by_date(self):
        self.transactions = sorted(self.transactions, key=lambda transaction: transaction.date)

    def add_commsec_transactions(self, file):
        transactions = []

        with open(file, "r") as f:
            csv_reader = csv.reader(f)
            next(csv_reader) # skip header row

            for row in csv_reader:
                details = row[2]
                details = details.split(" ")

                if details[0] == "Direct": continue # skip direct transfers (money deposits or withdrawals)

                date = row[0]
                date = datetime.strptime(date, "%d/%m/%Y")

                if details[0] == "B":
                    action = "buy"
                elif details[0] == "S":
                    action = "sell"

                stock = details[2]
                units = int(details[1])
                
                if action == "buy":
                    amount = float(row[3])
                elif action == "sell":
                    amount = float(row[4])

                transaction = Portfolio.Transaction(date, action, stock, units, amount)
                transactions.append(transaction)

        transactions = transactions[::-1]
        self.transactions += transactions
        self.sort_transactions_by_date()

    def add_stake_transactions(self, file):
        transactions = []
        wb = load_workbook("Stake_AccountSummaryAUS_010201-120724.xlsx")
        sheet = wb["Trades"]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date = row[1]
            date = datetime.strptime(date, "%d-%m-%Y")

            if row[3] == "BUY":
                action = "buy"
            elif row[3] == "SELL":
                action = "sell"

            stock = row[0]
            units = int(row[4])
            amount = abs(float(row[6]))

            transaction = Portfolio.Transaction(date, action, stock, units, amount)
            transactions.append(transaction)
        
        transactions = transactions[::-1]
        self.transactions += transactions
        self.sort_transactions_by_date()

    def add_dividends(self, file):
        with open(file, "r") as f:
            csv_reader = csv.reader(f)
            next(csv_reader) # skip header row

            for row in csv_reader:
                date = row[0]
                date = datetime.strptime(date, "%d/%m/%Y")

                type = row[2]

                if type == "Direct Credit":
                    amount = float(row[1])
                elif type == "DRP":
                    action = "buy"
                    stock = row[3]
                    units = int(row[5])

                    if units == 0: continue

                    drp_share_price = float(row[4])
                    amount = units * drp_share_price

                    transaction = Portfolio.Transaction(date, action, stock, units, amount)
                    self.transactions.append(transaction)

                self.dividends.append(Portfolio.Dividend(date, amount))

        self.sort_transactions_by_date()
