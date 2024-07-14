import csv
from datetime import datetime
from openpyxl import load_workbook

class Transactions:
    class Transaction:
        def __init__(self, date, action, stock, units, amount):
            self.date = date
            self.action = action
            self.stock = stock
            self.units = units
            self.amount = amount

    def __init__(self):
        self.transactions = []

    def print_transactions(self):
        for t in self.transactions:
            print(t.date, t.action, t.stock, t.units, t.amount)
            
    def sort_transactions_by_date(self):
        self.transactions = sorted(self.transactions, key=lambda transaction: transaction.date)

    def add_commsec_transactions(self, file):
        with open(file, "r") as f:
            csv_reader = csv.reader(f)
            next(csv_reader) # skip header row

            for row in csv_reader:
                details = row[2]
                details = details.split(" ")

                if details[0] == "Direct": continue # skip direct transfers (money deposits or withdrawals)

                date = row[0]
                date = datetime.strptime(date, "%d/%m/%Y")

                if details[0] == "B":
                    action = "buy"
                elif details[0] == "S":
                    action = "sell"

                stock = details[2]
                units = int(details[1])
                
                if action == "buy":
                    amount = float(row[3])
                elif action == "sell":
                    amount = float(row[4])

                transaction = Transactions.Transaction(date, action, stock, units, amount)
                self.transactions.append(transaction)

        self.sort_transactions_by_date()

    def add_stake_transactions(self, file):
        wb = load_workbook("Stake_AccountSummaryAUS_010201-120724.xlsx")
        sheet = wb["Trades"]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            date = row[1]
            date = datetime.strptime(date, "%d-%m-%Y")

            if row[3] == "BUY":
                action = "buy"
            elif row[3] == "SELL":
                action = "sell"

            stock = row[0]
            units = int(row[4])
            amount = abs(float(row[6]))

            transaction = Transactions.Transaction(date, action, stock, units, amount)
            self.transactions.append(transaction)
        
        self.sort_transactions_by_date()
